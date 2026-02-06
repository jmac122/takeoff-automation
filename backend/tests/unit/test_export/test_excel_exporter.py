"""Tests for Excel export generation."""

import time
import uuid

import pytest
from io import BytesIO
from openpyxl import load_workbook

from app.services.export.base import ExportData, ConditionData, MeasurementData
from app.services.export.excel_exporter import ExcelExporter


class TestExcelExporter:

    @pytest.fixture
    def exporter(self):
        return ExcelExporter()

    def test_content_type(self, exporter):
        """Content type is Excel OOXML."""
        assert exporter.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_file_extension(self, exporter):
        """File extension is .xlsx."""
        assert exporter.file_extension == ".xlsx"

    def test_generates_valid_excel(self, exporter, sample_project_data):
        """Output is a valid Excel workbook."""
        result = exporter.generate(sample_project_data)
        assert isinstance(result, bytes)
        wb = load_workbook(BytesIO(result))
        assert wb.sheetnames  # Has at least one sheet

    def test_summary_sheet_has_correct_totals(self, exporter, sample_project_data):
        """Summary sheet shows correct total quantity per condition."""
        result = exporter.generate(sample_project_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]

        # Header row is at row 4, data starts at row 5
        # Check Floor Slab total
        assert ws.cell(row=5, column=1).value == "Floor Slab"
        assert ws.cell(row=5, column=4).value == 1500.0
        assert ws.cell(row=5, column=5).value == 3

        # Check Footing
        assert ws.cell(row=6, column=1).value == "Footing"
        assert ws.cell(row=6, column=4).value == 250.0
        assert ws.cell(row=6, column=5).value == 2

        # Check Column Pads
        assert ws.cell(row=7, column=1).value == "Column Pads"
        assert ws.cell(row=7, column=4).value == 8.0
        assert ws.cell(row=7, column=5).value == 0

    def test_detail_sheet_per_condition(self, exporter, sample_project_data):
        """Each condition with measurements gets its own detail sheet."""
        result = exporter.generate(sample_project_data)
        wb = load_workbook(BytesIO(result))
        sheet_names = wb.sheetnames

        # Floor Slab and Footing have measurements, Column Pads does not
        assert "Floor Slab" in sheet_names
        assert "Footing" in sheet_names
        assert "Column Pads" not in sheet_names

    def test_empty_project_exports_cleanly(self, exporter, empty_project_data):
        """Project with no measurements produces valid Excel with summary only."""
        result = exporter.generate(empty_project_data)
        wb = load_workbook(BytesIO(result))
        assert "Summary" in wb.sheetnames
        # No condition detail sheets
        assert len(wb.sheetnames) == 1

    def test_special_characters_in_condition_names(self, exporter, sample_project_id):
        """Condition names with slashes, quotes, etc. don't break sheet names."""
        data = ExportData(
            project_id=sample_project_id,
            project_name="Test",
            project_description=None,
            client_name=None,
            conditions=[
                ConditionData(
                    id=uuid.uuid4(),
                    name="Slab [Level 1/2]: Main \"Area\"",
                    description=None,
                    scope="concrete",
                    category=None,
                    measurement_type="area",
                    color="#000000",
                    unit="SF",
                    depth=None,
                    thickness=None,
                    total_quantity=100.0,
                    measurement_count=1,
                    building=None,
                    area=None,
                    elevation=None,
                    measurements=[
                        MeasurementData(
                            id=uuid.uuid4(),
                            condition_name="Slab [Level 1/2]: Main \"Area\"",
                            condition_id=uuid.uuid4(),
                            page_id=uuid.uuid4(),
                            page_number=1,
                            sheet_number="S-001",
                            sheet_title=None,
                            geometry_type="polygon",
                            geometry_data={"points": [{"x": 0, "y": 0}, {"x": 10, "y": 0}, {"x": 10, "y": 10}]},
                            quantity=100.0,
                            unit="SF",
                            pixel_length=None,
                            pixel_area=100.0,
                            is_ai_generated=False,
                            is_verified=False,
                            notes=None,
                        ),
                    ],
                ),
            ],
        )
        result = exporter.generate(data)
        wb = load_workbook(BytesIO(result))
        # Should not raise and should have the sanitized sheet
        assert len(wb.sheetnames) >= 2

    def test_large_dataset_performance(self, exporter, sample_project_id):
        """Export with 1000+ measurements completes in under 10 seconds."""
        cond_id = uuid.uuid4()
        page_id = uuid.uuid4()
        measurements = [
            MeasurementData(
                id=uuid.uuid4(),
                condition_name="Slab",
                condition_id=cond_id,
                page_id=page_id,
                page_number=1,
                sheet_number="S-001",
                sheet_title="Plan",
                geometry_type="polygon",
                geometry_data={"points": [{"x": 0, "y": 0}, {"x": 10, "y": 0}, {"x": 10, "y": 10}]},
                quantity=float(i),
                unit="SF",
                pixel_length=None,
                pixel_area=100.0,
                is_ai_generated=False,
                is_verified=False,
                notes=None,
            )
            for i in range(1200)
        ]

        data = ExportData(
            project_id=sample_project_id,
            project_name="Large Project",
            project_description=None,
            client_name=None,
            conditions=[
                ConditionData(
                    id=cond_id,
                    name="Slab",
                    description=None,
                    scope="concrete",
                    category=None,
                    measurement_type="area",
                    color="#000000",
                    unit="SF",
                    depth=None,
                    thickness=None,
                    total_quantity=sum(range(1200)),
                    measurement_count=1200,
                    building=None,
                    area=None,
                    elevation=None,
                    measurements=measurements,
                ),
            ],
        )

        start = time.time()
        result = exporter.generate(data)
        elapsed = time.time() - start
        assert elapsed < 10.0, f"Export took {elapsed:.1f}s, expected < 10s"
        assert len(result) > 0

    def test_measurement_types_formatted_correctly(self, exporter, sample_project_data):
        """Area shows SF, linear shows LF, volume shows CY, count shows EA."""
        result = exporter.generate(sample_project_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]

        # Floor Slab unit is SF
        assert ws.cell(row=5, column=3).value == "SF"
        # Footing unit is LF
        assert ws.cell(row=6, column=3).value == "LF"
        # Column Pads unit is EA
        assert ws.cell(row=7, column=3).value == "EA"

    def test_page_references_included(self, exporter, sample_project_data):
        """Each measurement row includes the page/sheet number it came from."""
        result = exporter.generate(sample_project_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Floor Slab"]

        # Data starts at row 5 (row 4 is header)
        assert ws.cell(row=5, column=1).value == 1  # page_number
        assert ws.cell(row=5, column=2).value == "S-001"  # sheet_number

    def test_project_name_in_summary(self, exporter, sample_project_data):
        """Summary sheet contains the project name."""
        result = exporter.generate(sample_project_data)
        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]
        assert "Test Construction Project" in str(ws.cell(row=1, column=1).value)

    def test_by_page_sheet_created(self, exporter, sample_project_data):
        """A 'By Page' sheet is created grouping measurements by page."""
        result = exporter.generate(sample_project_data)
        wb = load_workbook(BytesIO(result))
        assert "By Page" in wb.sheetnames
