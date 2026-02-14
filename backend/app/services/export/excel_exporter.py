"""Excel (.xlsx) exporter using openpyxl."""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.services.export.base import BaseExporter, ExportData, ConditionData, format_unit, sanitize_field


# Excel sheet name constraints
_INVALID_SHEET_CHARS = re.compile(r'[\\/*?\[\]:]')
_MAX_SHEET_NAME_LEN = 31


def _sanitize_sheet_name(name: str) -> str:
    """Make a string safe for use as an Excel sheet name."""
    name = _INVALID_SHEET_CHARS.sub('_', name)
    if len(name) > _MAX_SHEET_NAME_LEN:
        name = name[:_MAX_SHEET_NAME_LEN - 1] + '…'
    return name or "Sheet"


class ExcelExporter(BaseExporter):
    """Export project data to Excel workbook."""

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    def generate(self, data: ExportData, options: dict | None = None) -> bytes:
        wb = Workbook()
        self._build_summary_sheet(wb, data)
        self._build_detail_sheets(wb, data)
        self._build_page_sheets(wb, data)
        self._build_cost_sheet(wb, data)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Styles
    # ------------------------------------------------------------------

    _HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    _HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    _HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _apply_header(self, ws, row: int, headers: list[str]) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = self._HEADER_ALIGN
            cell.border = self._THIN_BORDER

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------

    def _build_summary_sheet(self, wb: Workbook, data: ExportData) -> None:
        ws = wb.active
        ws.title = "Summary"

        # Title row
        ws.cell(row=1, column=1, value=f"Export: {sanitize_field(data.project_name)}").font = Font(bold=True, size=14)
        if data.client_name:
            ws.cell(row=2, column=1, value=f"Client: {sanitize_field(data.client_name)}")

        # Summary table
        has_costs = any(c.assembly_cost for c in data.conditions)
        headers = ["Condition", "Type", "Unit", "Quantity", "Measurements", "Category", "Scope"]
        if has_costs:
            headers += ["Unit Cost", "Material", "Labor", "Total Cost", "Markup Total"]
        start_row = 4
        self._apply_header(ws, start_row, headers)

        for idx, cond in enumerate(data.conditions):
            row = start_row + 1 + idx
            ws.cell(row=row, column=1, value=sanitize_field(cond.name))
            ws.cell(row=row, column=2, value=sanitize_field(cond.measurement_type))
            ws.cell(row=row, column=3, value=format_unit(cond.unit))
            ws.cell(row=row, column=4, value=cond.total_quantity)
            ws.cell(row=row, column=5, value=cond.measurement_count)
            ws.cell(row=row, column=6, value=sanitize_field(cond.category or ""))
            ws.cell(row=row, column=7, value=sanitize_field(cond.scope))
            if has_costs and cond.assembly_cost:
                ac = cond.assembly_cost
                ws.cell(row=row, column=8, value=ac.unit_cost)
                ws.cell(row=row, column=9, value=ac.material_cost)
                ws.cell(row=row, column=10, value=ac.labor_cost)
                ws.cell(row=row, column=11, value=ac.total_cost)
                ws.cell(row=row, column=12, value=ac.total_with_markup)

        # Project cost total row
        if has_costs:
            total_row = start_row + 1 + len(data.conditions)
            ws.cell(row=total_row, column=1, value="PROJECT TOTAL").font = Font(bold=True)
            ws.cell(row=total_row, column=12, value=data.total_project_cost).font = Font(bold=True)

        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ------------------------------------------------------------------
    # Per-condition detail sheets
    # ------------------------------------------------------------------

    def _build_detail_sheets(self, wb: Workbook, data: ExportData) -> None:
        seen_names: dict[str, int] = {}
        for cond in data.conditions:
            if not cond.measurements:
                continue

            base_name = _sanitize_sheet_name(cond.name)
            if base_name in seen_names:
                seen_names[base_name] += 1
                suffix = f" ({seen_names[base_name]})"
                # Truncate base to leave room for the counter within the 31-char budget
                sanitized = _INVALID_SHEET_CHARS.sub('_', cond.name)
                max_base_len = _MAX_SHEET_NAME_LEN - len(suffix)
                if len(sanitized) > max_base_len:
                    sanitized = sanitized[:max_base_len]
                sheet_name = sanitized + suffix
            else:
                seen_names[base_name] = 1
                sheet_name = base_name

            ws = wb.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value=sanitize_field(cond.name)).font = Font(bold=True, size=12)
            ws.cell(row=2, column=1, value=f"Type: {sanitize_field(cond.measurement_type)}  |  Unit: {format_unit(cond.unit)}  |  Total: {cond.total_quantity:.2f}")

            headers = ["Page", "Sheet #", "Geometry", "Quantity", "Unit", "Verified", "Notes"]
            self._apply_header(ws, 4, headers)

            for m_idx, m in enumerate(cond.measurements):
                row = 5 + m_idx
                ws.cell(row=row, column=1, value=m.page_number)
                ws.cell(row=row, column=2, value=sanitize_field(m.sheet_number or ""))
                ws.cell(row=row, column=3, value=m.geometry_type)
                ws.cell(row=row, column=4, value=m.quantity)
                ws.cell(row=row, column=5, value=format_unit(m.unit))
                ws.cell(row=row, column=6, value="Yes" if m.is_verified else "No")
                ws.cell(row=row, column=7, value=sanitize_field(m.notes or ""))

            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # ------------------------------------------------------------------
    # Cost summary sheet
    # ------------------------------------------------------------------

    def _build_cost_sheet(self, wb: Workbook, data: ExportData) -> None:
        costed = [c for c in data.conditions if c.assembly_cost]
        if not costed:
            return

        ws = wb.create_sheet(title="Cost Summary")
        ws.cell(row=1, column=1, value="Cost Summary").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Project: {sanitize_field(data.project_name)}")

        headers = [
            "Condition", "Qty", "Unit", "Unit Cost",
            "Material", "Labor", "Equipment", "Subcontract", "Other",
            "Total Cost", "OH %", "Profit %", "Total w/ Markup",
        ]
        self._apply_header(ws, 4, headers)

        for idx, cond in enumerate(costed):
            row = 5 + idx
            ac = cond.assembly_cost
            ws.cell(row=row, column=1, value=sanitize_field(cond.name))
            ws.cell(row=row, column=2, value=cond.total_quantity)
            ws.cell(row=row, column=3, value=format_unit(cond.unit))
            ws.cell(row=row, column=4, value=ac.unit_cost)
            ws.cell(row=row, column=5, value=ac.material_cost)
            ws.cell(row=row, column=6, value=ac.labor_cost)
            ws.cell(row=row, column=7, value=ac.equipment_cost)
            ws.cell(row=row, column=8, value=ac.subcontract_cost)
            ws.cell(row=row, column=9, value=ac.other_cost)
            ws.cell(row=row, column=10, value=ac.total_cost)
            ws.cell(row=row, column=11, value=ac.overhead_percent)
            ws.cell(row=row, column=12, value=ac.profit_percent)
            ws.cell(row=row, column=13, value=ac.total_with_markup)

        total_row = 5 + len(costed)
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for col, attr in [(5, "material_cost"), (6, "labor_cost"), (7, "equipment_cost"),
                          (8, "subcontract_cost"), (9, "other_cost"), (10, "total_cost"),
                          (13, "total_with_markup")]:
            ws.cell(row=total_row, column=col,
                    value=sum(getattr(c.assembly_cost, attr) for c in costed)).font = Font(bold=True)

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # ------------------------------------------------------------------
    # Per-page sheets
    # ------------------------------------------------------------------

    def _build_page_sheets(self, wb: Workbook, data: ExportData) -> None:
        # Group measurements by page
        pages: dict[int | None, list] = {}
        for m in data.all_measurements:
            pages.setdefault(m.page_number, []).append(m)

        if not pages:
            return

        ws = wb.create_sheet(title="By Page")
        headers = ["Page", "Sheet #", "Condition", "Geometry", "Quantity", "Unit"]
        self._apply_header(ws, 1, headers)

        current_row = 2
        for page_num in sorted(pages.keys(), key=lambda x: (x is None, x)):
            for m in pages[page_num]:
                ws.cell(row=current_row, column=1, value=m.page_number)
                ws.cell(row=current_row, column=2, value=sanitize_field(m.sheet_number or ""))
                ws.cell(row=current_row, column=3, value=sanitize_field(m.condition_name))
                ws.cell(row=current_row, column=4, value=m.geometry_type)
                ws.cell(row=current_row, column=5, value=m.quantity)
                ws.cell(row=current_row, column=6, value=format_unit(m.unit))
                current_row += 1

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
